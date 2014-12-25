import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "migrants.settings")
import django
django.setup()

from invoke import task
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter
import pycountry

from migrants.base.models import Country, DataCategory, MigrationInfo


try:
    _range = xrange  # py2 :(
except NameError:
    _range = range  # py3 :)


country_columns = ['id', 'alt_name', 'order', 'area']
country_columns = {index + 1: col for index, col in enumerate(country_columns)}
country_columns[7] = 'region'


def _countries_by_id():
    countries = {}

    for country in pycountry.countries:
        country_dict = {
            "alpha2": country.alpha2,
            "name": country.name,
        }
        countries[int(country.numeric)] = country_dict

    return countries


def _db_countries_by_name():
    countries = Country.objects.all()
    if not countries.exists():
        raise Exception("Load the countries first !")

    return {country.alt_name: country for country in countries}


def _get_worksheet(name, filename='data.xlsx'):
    wb = load_workbook(filename=filename, use_iterators=False)
    ws = wb.get_sheet_by_name(name=name)
    return ws


@task
def import_countries():
    countries = _countries_by_id()
    ws = _get_worksheet('ANNEX')

    to_insert = []
    for row in _range(16, 248):
        country_dict = {}

        for col_idx, name in country_columns.items():
            col = get_column_letter(col_idx)
            key = "{}{}".format(col, row)
            country_dict[name] = ws.cell(key).value

        try:
            pk = country_dict['id']
            country_dict.update(countries[pk])
            to_insert.append(Country(**country_dict))
        except KeyError:
            # Channel Islands
            print("Skipping country {} ...".format(country_dict['alt_name']))

    Country.objects.bulk_create(to_insert)


@task
def import_categories():
    ws = _get_worksheet('CONTENTS')
    categories = []

    for row in _range(17, 29):
        table_key = '{}{}'.format(get_column_letter(1), row)
        title_key = '{}{}'.format(get_column_letter(2), row)

        table = ws.cell(table_key).value
        title = ws.cell(title_key).value
        title, year = title[0:-7], int(title[-4:])

        pk = int(table.split(" ")[-1])

        category = DataCategory(id=pk, title=title, year=year)
        categories.append(category)

    DataCategory.objects.bulk_create(categories)


def import_category_country(ws, row, category, countries):
    result = []
    destination_name = ws.cell("B{}".format(row)).value
    try:
        destication_country = countries[destination_name]
    except KeyError:
        # This are the regions e.g Middle Africa (they have no data)
        print("Skippiing {} ...".format(destination_name))
        return

    for col_index in _range(10, 999):
        col = get_column_letter(col_index)
        if col == 'IH':
            break

        people = ws.cell("{}{}".format(col, row)).value
        if not people:
            continue
        try:
            people = int(people.replace(" ", ""))
        except AttributeError:
            # Allready an int, ignore it
            pass

        origin_country_name = ws.cell("{}16".format(col)).value
        try:
            origin_country = countries[origin_country_name]
        except KeyError:
            # Channel Islands
            print("Skipping country {} ....".format(origin_country_name))
            continue

        info = MigrationInfo(destination=destication_country,
                             origin=origin_country,
                             people=people,
                             category=category)
        result.append(info)

    MigrationInfo.objects.bulk_create(result)


def import_category(category, countries):
    sheet = 'Table {}'.format(category.id)
    ws = _get_worksheet(sheet)

    for row in _range(25, 282):
        import_category_country(ws, row, category, countries)


@task
def import_data():
    countries = _db_countries_by_name()
    for category in DataCategory.objects.all():
        import_category(category, countries)


@task
def import_all():
    import_countries()
    import_categories()
    import_data()
